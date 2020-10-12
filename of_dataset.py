#Nikesh Ghimire
#Summer Research 2020
#Optical Flow Dataset Maker

#Importing Modules
import cv2
import numpy as np
import json
import openpyxl as xl
import os 
import csv
import pandas
import shutil
from random import randrange

#Extracts frames from  video in name, with instName as name and location of video
#Places frames in either playing or n_playing folder depending on silences.json
def frameOut(path, instName):             

    try:

        for i in os.listdir(path):
            if i.endswith('.mp4'):# or i.endswith('.MOV'):
                vid = path + i
                break
        
        output_file = path + i[:i.rfind('.')] + '_op.avi'
        fourcc = cv2.VideoWriter_fourcc(*'DIVX')

        cap = cv2.VideoCapture(vid)
        ret, frame1 = cap.read()
        prvs = cv2.cvtColor(frame1,cv2.COLOR_BGR2GRAY)
        hsv = np.zeros_like(frame1)
        hsv[...,1] = 255

        frame_num = 0

        is_begin = True
    
    except Exception as e:

        print(e)
        pass

    try:
        while(1):
            ret, frame2 = cap.read()
            processed = frame2

            if is_begin:
                h, w, _ = processed.shape
                out = cv2.VideoWriter(output_file, fourcc, 30, (w, h), True)
                is_begin = False

            next = cv2.cvtColor(frame2,cv2.COLOR_BGR2GRAY)

            if frame2 is None:
                break

            flow = cv2.calcOpticalFlowFarneback(prvs,next, None, 0.5, 3, 15, 3, 5, 1.2, 0)
            

            mag, ang = cv2.cartToPolar(flow[...,0], flow[...,1])
            hsv[...,0] = ang*180/np.pi/2
            hsv[...,2] = cv2.normalize(mag,None,0,255,cv2.NORM_MINMAX)
            rgb = cv2.cvtColor(hsv,cv2.COLOR_HSV2BGR)

            # dense_flow = cv2.addWeighted(frame, 1,rgb, 2, 0)

            out.write(rgb)
            cv2.imshow('frame2',rgb)
            k = cv2.waitKey(30) & 0xff
            if k == 27:
                break
            elif k == ord('s'):
                cv2.imwrite('opticalfb.png',frame2)
                cv2.imwrite('opticalhsv.png',rgb)

            #Check playing or not playing from JSON
            
            s_list = compileInterval(path)
            playing = checkPlay(s_list, frame_num)

            dir_name = path[path[:-1].rfind('/') + 1:-1]

            if playing:
                cv2.imwrite('URMP/data/' + instName + '/playing/op_' + dir_name + '_' + str(frame_num) + '_' + instName + '.jpg',rgb)
            else:
                cv2.imwrite('URMP/data/' + instName + '/n_playing/op_' + dir_name + '_' + str(frame_num) + '_' + instName + '.jpg',rgb)
            frame_num = frame_num + 1
            prvs = next

        cap.release()
        cv2.destroyAllWindows()
    
    except Exception as e:
        print(e)
        pass

#HELPER FUNCTION
#Compiles silence intervals from the json files
def compileInterval(path):

    for i in os.listdir(path):
        if i.endswith('.JSON'):
            break

    with open(path + i, 'r', encoding='utf-8-sig') as inFile:
        data = json.load(inFile)

    s_list = [[round(eval(k),2), round(v,2)] for k, v in data['silences'].items()] 
    
    return s_list

#HELPER FUNCTION
#Returns boolean value of whether the current frame is playing or n_playing
def checkPlay(int_list, currFrame):

    currTime = currFrame/30

    for i in int_list:
        if currTime > i[0] and currTime < i[1]:
            return False
        else:
            pass
    
    return True

#Compiles a CSV for the folder path and instrument instName
def compileCSV(path, instName):

    instPath = path + '/' + instName + '/'

    n_playing = os.listdir(instPath + 'n_playing')
    playing = os.listdir(instPath + 'playing')
    test = os.listdir(instPath + 'test')

    wb = xl.Workbook()
    ws = wb.active

    ws.cell(row = 1, column = 1).value = 'name'
    ws.cell(row = 1, column = 2).value = 'label'

    for i in test:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'test/' + i
        ws.cell(row = row_num, column = 2).value = 'test' 

    for i in n_playing:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'n_playing/' + i
        ws.cell(row = row_num, column = 2).value = 'n_playing'

    for i in playing:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'playing/' + i
        ws.cell(row = row_num, column = 2).value = 'playing'

    wb.save(instPath + '/cleaned_pre.xlsx')

    data_xls = pandas.read_excel(instPath + '/cleaned_pre.xlsx', 'Sheet', index_col=None)
    data_xls.to_csv(instPath + '/cleaned.csv', encoding='utf-8', index=False)

    for i in os.listdir(instPath):
        if i.endswith('.xlsx'):
            os.remove(instPath + '/' + i)

#Resizes a dataset for an instrument based on percentage (for both p and np)
def resize(percent, path, instName):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    test = os.listdir(instPath + '/test')

    moved = 0

    pList = []
    npList = []

    for i in range(round(percent * len(n_playing))):
        
        r_num = randrange(0, len(n_playing))
        x = n_playing[r_num]
        while x in npList:
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
        npList.append(x)
        os.rename(instPath + '/n_playing/' + x, path + instName + '_bckup/bckup/n_playing/' + x)
        moved += 1

    for i in range(round(percent * len(playing))):
        r_num = randrange(0, len(playing))
        x = playing[r_num]
        while x in pList:
            r_num = randrange(0, len(playing))
            x = playing[r_num]
        pList.append(x)
        os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
        moved += 1

    print('Moved ' + str(moved) + ' files')

#Does the same as the one above, but only for either playing or n_playing (use the variables 'n' and 'p' for n_playing and playing respectfully)
def resize_one(percent, path, instName, pnp):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    test = os.listdir(instPath + '/test')

    moved = 0

    pList = []
    npList = []

    if pnp == 'n':
        for i in range(round(percent * len(n_playing))):
            
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
            while x in npList:
                r_num = randrange(0, len(n_playing))
                x = n_playing[r_num]
            npList.append(x)
            os.rename(instPath + '/n_playing/' + x, path + instName + '_bckup/bckup/n_playing/' + x)
            moved += 1
    elif pnp == 'p':
        for i in range(round(percent * len(playing))):
            r_num = randrange(0, len(playing))
            x = playing[r_num]
            while x in pList:
                r_num = randrange(0, len(playing))
                x = playing[r_num]
            pList.append(x)
            os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
            moved += 1

    print('Moved ' + str(moved) + ' files')

#Makes both datasets the same size 
def sameSize(path, instName):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    
    moved = 0

    print(len(playing), 'playing')
    print(len(n_playing), 'n_playing')

    pList = []
    nP = False
    P = False

    pBig = len(playing) - len(n_playing)
    nBig = len(n_playing) - len(playing)

    if nBig >= pBig:
        use = nBig
        for i in range(use):
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
            while x in pList:
                r_num = randrange(0, len(n_playing))
                x = n_playing[r_num]
            pList.append(x)
            os.rename(instPath + '/n_playing/' + x, path + instName + '_bckup/bckup/n_playing/' + x)
            moved += 1

        print('Resized n_playing w.r.t. playing --> Moved ' + str(moved) + ' files')
    else:
        use = pBig
        for i in range(use):
            
            r_num = randrange(0, len(playing))
            x = playing[r_num]
            while x in pList:
                r_num = randrange(0, len(playing))
                x = playing[r_num]
            pList.append(x)
            os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
            moved += 1

        print('Resized playing w.r.t. n_playing --> Moved ' + str(moved) + ' files')

#Moves files to a different directly (specify number of files)
def move(instName, path, n_path, num_files, pnp):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    
    moved = 0

    pList = [] 

    if pnp == 'n':
        for i in range(num_files):
            
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
            while x in pList:
                r_num = randrange(0, len(n_playing))
                x = n_playing[r_num]
            pList.append(x)
            os.rename(instPath + '/n_playing/' + x, n_path + '/n_playing/cello_moved_' + str(moved) + '.jpg')
            moved += 1

        print('Resized n_playing w.r.t. playing --> Moved ' + str(moved) + ' files')

    elif pnp == 'p':
        for i in range(num_files):
            
            r_num = randrange(0, len(playing))
            x = playing[r_num]
            while x in pList:
                r_num = randrange(0, len(playing))
                x = playing[r_num]
            pList.append(x)
            os.rename(instPath + '/playing/' + x, n_path + '/playing/cello_moved_' + str(moved) + '.jpg')
            moved += 1

        print('Resized playing w.r.t. n_playing --> Moved ' + str(moved) + ' files')
    
    else:
        print('pnp parameter is invalid')
    
#Runs frameOut function on all videos for a specific instrument. This gives use p or np frames in their respective folders 
def extract_flow(inst):

    try:
        os.mkdir('URMP/data/' + inst)
        os.mkdir('URMP/data/' + inst + '/playing')
        os.mkdir('URMP/data/' + inst + '/n_playing')
        os.mkdir('URMP/data/' + inst + '/test')
    except:
        pass  

    for i in os.listdir('cnn_set/' + inst +'/'):
        n_path = 'cnn_set/' + inst + '/' + i + '/'
        frameOut(n_path, inst)

    compileCSV('URMP/data/', inst)

def main():

    # for i in os.listdir('cnn_set/'):
    #     extract_flow(i)

    extract_flow('guitar_2_niko')

    # print('working')

    # frameOut('cnn_set/trombone/07_GString_tpt_tbn/', 'trombone')

    # resize_one(0.15, 'URMP/data/', 'violin_cello_vNN', 'n')
    # sameSize('URMP/data/', 'trumpet')
    # move('cello', 'URMP/data/', 'URMP/data/violin_cello_vNN/', 4385, 'n_playing')
    # compileCSV('URMP/data/', 'guitar_2_niko')
    

        

if __name__ == '__main__':

    main()